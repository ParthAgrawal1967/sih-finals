import os
import json
from typing import List, Optional
from datetime import datetime as dt_datetime
from urllib.request import urlopen
from urllib.error import URLError, HTTPError
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
import numpy as np
import pandas as pd
import joblib
import xgboost as xgb
from tensorflow.keras.models import load_model
from dotenv import load_dotenv
from overview_simulator.simulator import run_simulation as overview_run_simulation
from overview_simulator.models import (
    OverviewSimulateRequest,
    OverviewSimulateResponse,
)



load_dotenv()  



# ----------------------------
# CONFIG
# ----------------------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Find models and data directories with multiple fallbacks
# Try in order: backend/models, ../models, ../../models (for Render's /opt structure)
possible_model_roots = [
    os.path.join(BASE_DIR, "models"),
    os.path.join(BASE_DIR, "..", "models"),
    os.path.join(BASE_DIR, "..", "..", "models"),
]
possible_data_roots = [
    os.path.join(BASE_DIR, "data"),
    os.path.join(BASE_DIR, "..", "data"),
    os.path.join(BASE_DIR, "..", "..", "data"),
]

MODEL_ROOT = None
DATA_ROOT = None

for path in possible_model_roots:
    if os.path.exists(path):
        MODEL_ROOT = path
        break
if MODEL_ROOT is None:
    MODEL_ROOT = possible_model_roots[0]  # Use first as default even if not found

for path in possible_data_roots:
    if os.path.exists(path):
        DATA_ROOT = path
        break
if DATA_ROOT is None:
    DATA_ROOT = possible_data_roots[0]  # Use first as default even if not found

# Support both folder layouts:
# 1) backend/models/global/... and backend/models/indian/...
# 2) backend/models/cpo_lstm_model.h5, data_scaler.pkl, xgb_gen11_imports_tonnes.json
if os.path.exists(os.path.join(MODEL_ROOT, "global", "cpo_lstm_model.h5")):
    GLOBAL_MODEL_PATH = os.path.join(MODEL_ROOT, "global", "cpo_lstm_model.h5")
    GLOBAL_SCALER_PATH = os.path.join(MODEL_ROOT, "global", "data_scaler.pkl")
    INDIA_IMPORT_MODEL_PATH = os.path.join(MODEL_ROOT, "indian", "xgb_gen11_imports_tonnes.json")
elif os.path.exists(os.path.join(MODEL_ROOT, "cpo_lstm_model.h5")):
    GLOBAL_MODEL_PATH = os.path.join(MODEL_ROOT, "cpo_lstm_model.h5")
    GLOBAL_SCALER_PATH = os.path.join(MODEL_ROOT, "data_scaler.pkl")
    INDIA_IMPORT_MODEL_PATH = os.path.join(MODEL_ROOT, "xgb_gen11_imports_tonnes.json")
else:
    GLOBAL_MODEL_PATH = os.path.join(MODEL_ROOT, "global", "cpo_lstm_model.h5")
    GLOBAL_SCALER_PATH = os.path.join(MODEL_ROOT, "global", "data_scaler.pkl")
    INDIA_IMPORT_MODEL_PATH = os.path.join(MODEL_ROOT, "indian", "xgb_gen11_imports_tonnes.json")

INDIA_BASE_FILE = os.path.join(DATA_ROOT, "india_cpo_clean_ml_dataset_gen11_with_landed.csv")
GLOBAL_DATA_FILE = os.path.join(DATA_ROOT, "global_dataset.csv")

TIME_STEPS = 3

# Economic constants
SWS_RATE = 10
IGST_RATE = 5
TONNE_TO_KG = 1000
CPO_DENSITY = 0.91
CFPI_WEIGHT = 0.028535

app = FastAPI(title="Palm Tariff Simulator")

from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # allow all (for now)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ----------------------------
# NMEO-OP statewise data
# ----------------------------
DEFAULT_NMEO_REGIONAL_DATA = [
    {
        "state": "Andhra Pradesh",
        "progress": 85,
        "palmArea": 1.8464,
        "farmerParticipation": 40200,
        "fundingUtilization": 78,
        "fundingAmount": 890,
        "processingUnits": 8,
        "achievements": [
            "Largest reported oil palm area among implementation states",
            "Strong processing ecosystem relative to other states",
            "Consistent mission participation under NMEO-OP"
        ],
        "challenges": [
            "Need to sustain FFB evacuation and processing logistics",
            "Further productivity gains in new plantations"
        ]
    },
    {
        "state": "Karnataka",
        "progress": 72,
        "palmArea": 0.46954,
        "farmerParticipation": 12300,
        "fundingUtilization": 65,
        "fundingAmount": 520,
        "processingUnits": 5,
        "achievements": [
            "Substantial area scale-up under mission period",
            "Improved extension and agronomy support",
            "Progressive adoption in suitable districts"
        ],
        "challenges": [
            "Water management and irrigation reliability",
            "Cost pressures for small and medium growers"
        ]
    },
    {
        "state": "Tamil Nadu",
        "progress": 68,
        "palmArea": 0.32982,
        "farmerParticipation": 9800,
        "fundingUtilization": 58,
        "fundingAmount": 380,
        "processingUnits": 3,
        "achievements": [
            "Steady area development in mission districts",
            "Good extension linkages for growers",
            "Improved planting material adoption"
        ],
        "challenges": [
            "Weather variability and moisture stress in pockets",
            "Need for stronger processing linkage in expansion zones"
        ]
    },
    {
        "state": "Mizoram",
        "progress": 64,
        "palmArea": 0.2668,
        "farmerParticipation": 7600,
        "fundingUtilization": 52,
        "fundingAmount": 290,
        "processingUnits": 2,
        "achievements": [
            "Strong progress among north-eastern implementation states",
            "Consistent area addition under mission support",
            "Improved farmer engagement in identified clusters"
        ],
        "challenges": [
            "Hilly terrain logistics and connectivity constraints",
            "Need for local processing and aggregation capacity"
        ]
    },
    {
        "state": "Odisha",
        "progress": 58,
        "palmArea": 0.2313,
        "farmerParticipation": 6900,
        "fundingUtilization": 45,
        "fundingAmount": 180,
        "processingUnits": 1,
        "achievements": [
            "Area expansion progressing in focused districts",
            "Improved field-level extension and training",
            "Better awareness for long-gestation crop planning"
        ],
        "challenges": [
            "Infrastructure gaps in remote plantation pockets",
            "Need to scale post-harvest and transport networks"
        ]
    },
    {
        "state": "Telangana",
        "progress": 45,
        "palmArea": 0.21382,
        "farmerParticipation": 6400,
        "fundingUtilization": 38,
        "fundingAmount": 220,
        "processingUnits": 1,
        "achievements": [
            "Early-stage area expansion reported in official records",
            "Strengthening mission implementation systems",
            "Gradual increase in farmer participation"
        ],
        "challenges": [
            "Need to accelerate area expansion trajectory",
            "Processing and logistics capacity must scale with area"
        ]
    }
]

NMEO_STATEWISE_SOURCE_URL = os.getenv("NMEO_STATEWISE_SOURCE_URL", "").strip()
NMEO_STATEWISE_SOURCE_NAME = os.getenv(
    "NMEO_STATEWISE_SOURCE_NAME",
    "Government of India Open Data (NMEO-OP state-wise dataset)"
)


def _to_float(value):
    try:
        if value is None:
            return None
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except Exception:
        return None


def _extract_lakh_ha(record):
    candidate_keys_lakh = [
        "palmArea", "palm_area_lakh_ha", "area_lakh_ha", "achievement_lakh_ha", "ach_lakh_ha"
    ]
    for key in candidate_keys_lakh:
        val = _to_float(record.get(key))
        if val is not None:
            return val

    latest_ach = _to_float(record.get("_2023_24___ach"))
    if latest_ach is not None:
        return latest_ach / 100000.0

    fallback_ach = _to_float(record.get("_2022_23___ach"))
    if fallback_ach is not None:
        return fallback_ach / 100000.0

    fallback_ach = _to_float(record.get("_2021_22___ach"))
    if fallback_ach is not None:
        return fallback_ach / 100000.0

    candidate_keys_ha = [
        "area_ha", "achievement_ha", "ach_ha", "area", "achievement", "ach"
    ]
    for key in candidate_keys_ha:
        val = _to_float(record.get(key))
        if val is not None:
            return val / 100000.0
    return None


def _extract_state(record):
    for key in ["state", "state_name", "state_ut", "state_name_ut"]:
        value = record.get(key)
        if value:
            return str(value).strip()
    return None


def _merge_nmeo_statewise(default_rows, external_rows):
    if not external_rows:
        return default_rows

    area_map = {}
    for row in external_rows:
        if not isinstance(row, dict):
            continue
        state = _extract_state(row)
        area = _extract_lakh_ha(row)
        if state and area is not None:
            area_map[state.lower()] = area

    merged = []
    for row in default_rows:
        row_copy = dict(row)
        state_key = row_copy["state"].lower()
        if state_key in area_map:
            row_copy["palmArea"] = round(area_map[state_key], 5)
        merged.append(row_copy)
    return merged


def _fetch_external_nmeo_rows():
    if not NMEO_STATEWISE_SOURCE_URL:
        return None
    try:
        with urlopen(NMEO_STATEWISE_SOURCE_URL, timeout=12) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
            if isinstance(payload, list):
                return payload
            if isinstance(payload, dict):
                if isinstance(payload.get("records"), list):
                    return payload["records"]
                if isinstance(payload.get("data"), list):
                    return payload["data"]
            return None
    except (HTTPError, URLError, TimeoutError, ValueError):
        return None


@app.get("/nmeo-op/statewise")
async def get_nmeo_statewise():
    external_rows = _fetch_external_nmeo_rows()
    regional_data = _merge_nmeo_statewise(DEFAULT_NMEO_REGIONAL_DATA, external_rows)
    return {
        "source": NMEO_STATEWISE_SOURCE_NAME,
        "source_url": NMEO_STATEWISE_SOURCE_URL or None,
        "as_of": dt_datetime.utcnow().strftime("%Y-%m-%d"),
        "regionalData": regional_data
    }


# ----------------------------
# Request / Response models
# ----------------------------
class SimulateRequest(BaseModel):
    tariff_pct: float
    horizon_months: int
    farmer_margin_pct: float
    last_global_window: Optional[List] = None
    scenario_type: Optional[str] = None
    scenario_parameters: Optional[dict] = None
    
    class Config:
        extra = "allow"  # Allow extra fields from frontend

class SimulateResponse(BaseModel):
    scenario: dict
    time_series: List[dict]
    final_summary: dict

# ----------------------------
# Load models / data at startup
# ----------------------------
STARTUP_ERROR = None
GLOBAL_MODEL = None
SCALER_X = None
SCALER_Y = None
FEATURE_NAMES = None
INDIAN_BST = None
DF_INDIA = None
LAST_INDIA_ROW = None
DF_GLOBAL = None

try:
    # Log paths for debugging
    print(f"DEBUG: GLOBAL_MODEL_PATH = {GLOBAL_MODEL_PATH}")
    print(f"DEBUG: GLOBAL_SCALER_PATH = {GLOBAL_SCALER_PATH}")
    print(f"DEBUG: INDIA_IMPORT_MODEL_PATH = {INDIA_IMPORT_MODEL_PATH}")
    print(f"DEBUG: INDIA_BASE_FILE = {INDIA_BASE_FILE}")
    print(f"DEBUG: GLOBAL_DATA_FILE = {GLOBAL_DATA_FILE}")
    
    GLOBAL_MODEL = load_model(GLOBAL_MODEL_PATH)
    scalers = joblib.load(GLOBAL_SCALER_PATH)
    SCALER_X = scalers["scaler_x"]
    SCALER_Y = scalers["scaler_y"]
    FEATURE_NAMES = list(scalers["feature_names"])

    INDIAN_BST = xgb.Booster()
    INDIAN_BST.load_model(INDIA_IMPORT_MODEL_PATH)

    DF_INDIA = pd.read_csv(INDIA_BASE_FILE).sort_values(["year", "month"]).reset_index(drop=True)
    LAST_INDIA_ROW = DF_INDIA.iloc[-1]

    if os.path.exists(GLOBAL_DATA_FILE):
        DF_GLOBAL = pd.read_csv(GLOBAL_DATA_FILE)
        if "date" in DF_GLOBAL.columns:
            DF_GLOBAL["date"] = pd.to_datetime(DF_GLOBAL["date"])
            DF_GLOBAL = DF_GLOBAL.sort_values("date").reset_index(drop=True)
    
    print("DEBUG: All models and data loaded successfully")
except Exception as e:
    STARTUP_ERROR = f"{type(e).__name__}: {str(e)}"
    print(f"ERROR during startup: {STARTUP_ERROR}")
    import traceback
    traceback.print_exc()

INDIAN_FEATURE_ORDER = [
    "year",
    "month",
    "tariff_pct",
    "tariff_change_pct",
    "tariff_shock",
    "tariff_3m_avg",
    "tariff_6m_avg",
    "global_cpo_price_usd_per_tonne",
    "usd_inr",
    "freight_usd",
    "domestic_consumption_tonnes",
    "domestic_production_tonnes",
    "demand_supply_gap",
    "imports_tonnes_lag1",
    "imports_tonnes_lag3",
    "imports_tonnes_lag6",
]

INDIAN_FEATURE_TEMPLATE = {
    "year": 2025,
    "month": 1,
    "tariff_pct": 0.0,
    "tariff_change_pct": 0.0,
    "tariff_shock": 0,
    "tariff_3m_avg": 0.0,
    "tariff_6m_avg": 0.0,
    "global_cpo_price_usd_per_tonne": 0.0,
    "usd_inr": 82.5,
    "freight_usd": 22.0,
    "domestic_consumption_tonnes": 950000.0,
    "domestic_production_tonnes": 400000.0,
    "demand_supply_gap": 550000.0,
    "imports_tonnes_lag1": 400000.0,
    "imports_tonnes_lag3": 420000.0,
    "imports_tonnes_lag6": 450000.0,
}

# ----------------------------
# Helpers: global forecast
# ----------------------------
def expand_price_to_feature_row(price, feature_names):
    if DF_GLOBAL is not None:
        try:
            last_row = DF_GLOBAL[feature_names].iloc[-1].values.astype(float)
            last_row[0] = float(price)
            return last_row
        except Exception:
            pass
    arr = np.zeros(len(feature_names), dtype=float)
    arr[0] = float(price)
    return arr

def recursive_global_forecast(last_window_input: List, months: int) -> List[float]:
    if GLOBAL_MODEL is None or SCALER_X is None or SCALER_Y is None:
        raise RuntimeError(f"Global model/scalers not loaded: {STARTUP_ERROR}")

    recent_window = []

    for item in last_window_input:
        if isinstance(item, (list, tuple, np.ndarray)) and len(item) == len(FEATURE_NAMES):
            recent_window.append(np.array(item, dtype=float))
            continue

        if isinstance(item, (list, tuple)) and len(item) == 1:
            item = item[0]

        try:
            price = float(item)
        except Exception:
            raise ValueError(f"Invalid value in last_global_window: {item}")

        recent_window.append(expand_price_to_feature_row(price, FEATURE_NAMES))

    while len(recent_window) < TIME_STEPS:
        recent_window.insert(0, recent_window[0].copy())
    recent_window = np.vstack(recent_window[-TIME_STEPS:])

    preds = []
    for _ in range(months):
        scaled_input = SCALER_X.transform(recent_window)
        reshaped = scaled_input.reshape(1, TIME_STEPS, scaled_input.shape[1])
        y_scaled = GLOBAL_MODEL.predict(reshaped)
        y = float(SCALER_Y.inverse_transform(y_scaled.reshape(-1, 1))[0][0])
        preds.append(y)

        new_row = recent_window[-1].copy()
        new_row[0] = y
        recent_window = np.vstack([recent_window[1:], new_row])

    return preds

# ----------------------------
# XGB predict wrapper
# ----------------------------
def predict_imports_from_xgb(X_df: pd.DataFrame) -> np.ndarray:
    dm = xgb.DMatrix(X_df)
    preds = INDIAN_BST.predict(dm)
    return preds

# ----------------------------
# Economics helpers
# ----------------------------
def compute_cif(global_price, usd_inr, freight_usd):
    return global_price * usd_inr + freight_usd

def compute_landed(cif, tariff_pct):
    return cif * (1 + (tariff_pct + SWS_RATE) / 100.0) * (1 + IGST_RATE / 100.0)

def compute_retail_price(landed_inr_per_tonne):
    return (landed_inr_per_tonne / (TONNE_TO_KG * CPO_DENSITY)) * 1.18

def compute_farmer_price(landed_inr_per_tonne, farmer_margin_fraction):
    return landed_inr_per_tonne * (1 - farmer_margin_fraction)

def compute_govt_revenue(imports_tonnes, landed, cif):
    return imports_tonnes * (landed - cif)

def compute_import_dependency(imports_tonnes, domestic_consumption_tonnes):
    if domestic_consumption_tonnes <= 0:
        return 0.0
    return (imports_tonnes / domestic_consumption_tonnes) * 100.0

# ----------------------------
# Main simulate endpoint
# ----------------------------
@app.post("/simulate", response_model=SimulateResponse)
def simulate(req: SimulateRequest):
    if STARTUP_ERROR:
        raise HTTPException(status_code=500, detail=f"Startup error: {STARTUP_ERROR}")

    tariff_pct = req.tariff_pct
    months = req.horizon_months
    farmer_margin_fraction = req.farmer_margin_pct / 100.0

    if req.last_global_window:
        last_window = req.last_global_window
    else:
        if DF_GLOBAL is not None:
            last_window = DF_GLOBAL[FEATURE_NAMES].iloc[-TIME_STEPS:].values.tolist()
        else:
            raise HTTPException(status_code=400, detail="No last_global_window provided and global data unavailable.")

    try:
        global_preds = recursive_global_forecast(last_window, months)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Global forecasting failed: {e}")

    base = LAST_INDIA_ROW
    imports_lag1 = float(base["imports_tonnes"]) if "imports_tonnes" in base else float(INDIAN_FEATURE_TEMPLATE["imports_tonnes_lag1"])
    imports_lag3 = imports_lag1
    imports_lag6 = imports_lag1

    last_tariff = float(base["tariff_pct"]) if "tariff_pct" in base else tariff_pct
    tariff_history = [last_tariff] * 6

    time_series = []
    for i, gp in enumerate(global_preds):
        row = INDIAN_FEATURE_TEMPLATE.copy()
        row["year"] = int(base.get("year", row["year"]))
        row["month"] = int(base.get("month", row["month"])) + i + 1

        row["global_cpo_price_usd_per_tonne"] = float(gp)
        row["usd_inr"] = float(base.get("usd_inr", row["usd_inr"]))
        row["freight_usd"] = float(base.get("freight_usd", row["freight_usd"]))
        row["domestic_consumption_tonnes"] = float(base.get("domestic_consumption_tonnes", row["domestic_consumption_tonnes"]))
        row["domestic_production_tonnes"] = float(base.get("domestic_production_tonnes", row["domestic_production_tonnes"]))
        row["demand_supply_gap"] = row["domestic_consumption_tonnes"] - row["domestic_production_tonnes"]

        row["tariff_pct"] = tariff_pct
        row["tariff_change_pct"] = float(tariff_pct - tariff_history[-1])
        row["tariff_shock"] = 1 if abs(row["tariff_change_pct"]) >= 5 else 0
        row["tariff_3m_avg"] = float(np.mean(tariff_history[-3:]))
        row["tariff_6m_avg"] = float(np.mean(tariff_history[-6:]))

        row["imports_tonnes_lag1"] = imports_lag1
        row["imports_tonnes_lag3"] = imports_lag3
        row["imports_tonnes_lag6"] = imports_lag6

        X_row = pd.DataFrame([[row[col] for col in INDIAN_FEATURE_ORDER]], columns=INDIAN_FEATURE_ORDER)

        try:
            pred_imports = float(predict_imports_from_xgb(X_row)[0])
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Indian model prediction error: {e}")

        imports_lag6 = imports_lag3
        imports_lag3 = imports_lag1
        imports_lag1 = pred_imports
        tariff_history.append(tariff_pct)

        cif = compute_cif(row["global_cpo_price_usd_per_tonne"], row["usd_inr"], row["freight_usd"])
        landed = compute_landed(cif, tariff_pct)
        retail = compute_retail_price(landed)
        farmer_price = compute_farmer_price(landed, farmer_margin_fraction)
        govt_rev = compute_govt_revenue(pred_imports, landed, cif)
        imp_dep_pct = compute_import_dependency(pred_imports, row["domestic_consumption_tonnes"])

        time_series.append({
            "month_index": i + 1,
            "global_price_usd_per_tonne": float(gp),
            "predicted_imports_tonnes": pred_imports,
            "landed_cost_inr_per_tonne": float(landed),
            "retail_price_inr_per_litre": float(retail),
            "farmer_price_inr_per_tonne": float(farmer_price),
            "government_revenue_inr": float(govt_rev),
            "import_dependency_pct": float(imp_dep_pct)
        })

    final_summary = time_series[-1] if len(time_series) > 0 else {}

    return {
        "scenario": {
            "tariff_pct": tariff_pct,
            "horizon_months": months,
            "farmer_margin_pct": req.farmer_margin_pct
        },
        "time_series": time_series,
        "final_summary": final_summary
    }

# ---------------------------------------------------------
# AI INTERPRETATION
# ---------------------------------------------------------
from huggingface_hub import InferenceClient

HF_API_KEY = os.getenv("HF_API_KEY")

client = InferenceClient(
    "meta-llama/Llama-3.1-70B-Instruct",
    token=HF_API_KEY
)

class InterpretRequest(BaseModel):
    tariff_change_pct: float
    summary: dict
    monthly_outputs: list

class InterpretResponse(BaseModel):
    overview: str
    economic_impact: str
    policy_recommendations: list
    risks: list


@app.post("/interpret", response_model=InterpretResponse)
async def interpret(request: InterpretRequest):

    tariff = request.tariff_change_pct
    summary = request.summary
    monthly = request.monthly_outputs

    prompt = f"""
You are an economic policy analyst specializing in agri-trade and tariff modelling.

### Simulation Inputs
- Tariff Change: {tariff}%
- Forecast Horizon: {len(monthly)} months

### Final Summary
{summary}

### Monthly Outputs
{monthly}

---

## TASKS

### 1. Narrative overview (5-7 sentences)
Explain impacts on imports, landed cost, domestic price, farmer income, and government revenue.

### 2. Bullet-point economic impact summary (use up/down arrows)

### 3. Policy recommendations (3-5)

### 4. Key risks (3)

### 5. ASCII Mini-Charts (use block characters)
Charts for:
- Predicted Monthly Imports  
- Landed Cost Trend  
- Import Dependency Trend  

IMPORTANT:
Output must be Markdown  
Do NOT invent numbers  
"""

    try:
        response_text = client.chat_completion(
            model="meta-llama/Llama-3.1-70B-Instruct",
            messages=[
                {"role": "system", "content": "You are an expert economic analyst."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.4,
        ).choices[0].message["content"]

    except Exception as e:
        print("HF API ERROR:", e)
        response_text = f"AI Interpretation failed: {e}"

    return {
        "overview": response_text,
        "economic_impact": "",
        "policy_recommendations": [],
        "risks": []
    }


# ---------- SIMULATE SCENARIO ENDPOINT ----------
from simulation.engine import simulate_scenario
from simulation.state import make_initial_state

class ScenarioSimRequest(BaseModel):
    tariff_pct: float
    horizon_months: int
    farmer_margin_pct: float
    scenario_type: Optional[str] = "baseline"
    scenario_parameters: Optional[dict] = None
    last_global_window: Optional[list] = None

class ScenarioSimResponse(BaseModel):
    time_series: List[dict]
    final_summary: dict

@app.post("/simulate_scenario", response_model=ScenarioSimResponse)
def simulate_scenario_endpoint(req: ScenarioSimRequest):
    if STARTUP_ERROR:
        raise HTTPException(status_code=500, detail=f"Startup error: {STARTUP_ERROR}")

    if req.last_global_window and len(req.last_global_window) > 0:
        last_price = None
        v = req.last_global_window[-1]
        if isinstance(v, (list, tuple)):
            last_price = float(v[0])
        else:
            try:
                last_price = float(v)
            except Exception:
                last_price = float(LAST_INDIA_ROW.get("global_cpo_price_usd_per_tonne", 0.0))
    else:
        if DF_GLOBAL is not None and "global_cpo_price_usd_per_tonne" in DF_GLOBAL.columns:
            last_price = float(DF_GLOBAL["global_cpo_price_usd_per_tonne"].iloc[-1])
        else:
            last_price = float(LAST_INDIA_ROW.get("global_cpo_price_usd_per_tonne", LAST_INDIA_ROW.get("global_price_usd_per_tonne", 0.0)))

    initial_state = make_initial_state(LAST_INDIA_ROW, last_price)

    try:
        out = simulate_scenario(
            initial_state=initial_state,
            horizon_months=int(req.horizon_months),
            scenario_type=req.scenario_type or "baseline",
            scenario_parameters=(req.scenario_parameters or {}),
            forecast_global_fn=recursive_global_forecast,
            predict_imports_fn=predict_imports_from_xgb,
            compute_cif_fn=compute_cif,
            compute_landed_fn=compute_landed,
            compute_retail_price_fn=compute_retail_price,
            compute_farmer_price_fn=compute_farmer_price,
            compute_govt_revenue_fn=compute_govt_revenue,
            compute_import_dependency_fn=compute_import_dependency,
            indian_feature_template=INDIAN_FEATURE_TEMPLATE,
            indian_feature_order=INDIAN_FEATURE_ORDER
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scenario simulation failed: {e}")

    return out

# ------------------------------------------------------------
# EXPORT ENDPOINTS (PDF + EXCEL)
# ------------------------------------------------------------
from fastapi import Response
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import datetime


class ExportRequest(BaseModel):
    tariffChange: float
    timeHorizon: int
    scenarioType: str
    scenarioParams: dict
    simulationData: dict
    interpretation: str


@app.post("/export/pdf")
async def export_pdf(req: ExportRequest):
    buffer = BytesIO()

    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    x = 40
    y = height - 50

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(x, y, "PalmTariff-AI - Government Policy Analytics")
    y -= 20

    pdf.setFont("Helvetica", 10)
    pdf.drawString(x, y, f"Export Timestamp: {datetime.datetime.now()}")
    y -= 30

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(x, y, "Simulation Parameters")
    y -= 20

    pdf.setFont("Helvetica", 10)
    pdf.drawString(x, y, f"Tariff Change: {req.tariffChange}%")
    y -= 15
    pdf.drawString(x, y, f"Time Horizon: {req.timeHorizon} months")
    y -= 15
    pdf.drawString(x, y, f"Scenario: {req.scenarioType}")
    y -= 15
    pdf.drawString(x, y, f"Parameters: {req.scenarioParams}")
    y -= 30

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(x, y, "KPI Summary")
    y -= 20

    summary = req.simulationData.get("final_summary", {})

    pdf.setFont("Helvetica", 10)
    for k, v in summary.items():
        pdf.drawString(x, y, f"{k}: {v}")
        y -= 15
        if y < 60:
            pdf.showPage()
            y = height - 50

    y -= 20

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(x, y, "AI Interpretation")
    y -= 20

    pdf.setFont("Helvetica", 10)
    for line in req.interpretation.split("\n"):
        pdf.drawString(x, y, line[:120])
        y -= 15
        if y < 60:
            pdf.showPage()
            y = height - 50

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return Response(
        content=buffer.read(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=tariff_simulation_report.pdf"
        }
    )


@app.post("/export/excel")
async def export_excel(req: ExportRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "Simulation Report"

    ws["A1"] = "PalmTariff-AI - Export Report"
    ws["A2"] = f"Timestamp: {datetime.datetime.now()}"

    ws.append([""])
    ws.append(["Simulation Parameters"])
    ws.append(["Tariff Change (%)", req.tariffChange])
    ws.append(["Time Horizon (months)", req.timeHorizon])
    ws.append(["Scenario Type", req.scenarioType])
    ws.append(["Scenario Params", str(req.scenarioParams)])

    ws.append([""])
    ws.append(["KPI Summary"])

    summary = req.simulationData.get("final_summary", {})
    for k, v in summary.items():
        ws.append([k, v])

    ws.append([""])
    ws.append(["Monthly Outputs"])

    monthly = req.simulationData.get("monthly_outputs", [])
    if monthly:
        headers = list(monthly[0].keys())
        ws.append(headers)
        for row in monthly:
            ws.append(list(row.values()))

    ws.append([""])
    ws.append(["AI Interpretation"])
    for line in req.interpretation.split("\n"):
        ws.append([line])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return Response(
        content=file_stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=tariff_simulation_report.xlsx"
        }
    )

@app.post("/overview/simulate", response_model=OverviewSimulateResponse)
async def overview_simulate(req: OverviewSimulateRequest):
    timeline = overview_run_simulation(
        horizon=req.horizon,
        actions=[a.dict(exclude_none=True) for a in req.actions],
        seed=req.seed,
    )
    return OverviewSimulateResponse(timeline=timeline)

from huggingface_hub import InferenceClient
HF_API_KEY = os.getenv("HF_API_KEY")

hf_client = InferenceClient(
    model="meta-llama/Llama-3.1-70B-Instruct",
    token=HF_API_KEY
)

class PolicyQuestionRequest(BaseModel):
    timeline: List[dict]
    question: str

class PolicyQuestionResponse(BaseModel):
    answer: str

@app.post("/overview/ask", response_model=PolicyQuestionResponse)
async def ask_policy_question(req: PolicyQuestionRequest):
    prompt = f"""
You are a senior trade and agriculture policy advisor to the Government of India.

Below is a deterministic simulation timeline for the global palm oil market:

{req.timeline}

QUESTION:
"{req.question}"

INSTRUCTIONS:
- Answer strictly based on the simulation context
- If exact numbers are unavailable, reason qualitatively
- Structure the answer under:
  1. Impact on Indian farmers
  2. Impact on Indian consumers
  3. Impact on government revenue
  4. Strategic trade implications
- Do NOT invent data
- Keep tone formal and policy-ready
"""

    try:
        response = hf_client.chat_completion(
            messages=[
                {"role": "system", "content": "You are an expert economic policy analyst."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=700,
            temperature=0.3,
        )

        answer_text = response.choices[0].message["content"]

    except Exception as e:
        answer_text = f"AI policy analysis failed: {e}"

    return {"answer": answer_text}


@app.get("/")
def root():
    return {"status": "ok", "startup_error": STARTUP_ERROR}